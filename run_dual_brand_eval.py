"""
run_dual_brand_eval.py
Run all Dobyns + Phenix eval questions through the live RAG pipeline
and print a pass/fail report. Uses the exact same system prompt as app_v3.py.
"""

import os, sys, re, time
from pathlib import Path

# ── Load .env ─────────────────────────────────────────────────────────────────
env_path = Path(__file__).parent / ".env"
for line in env_path.read_text().splitlines():
    line = line.strip()
    if "=" in line and not line.startswith("#"):
        k, v = line.split("=", 1)
        os.environ[k.strip()] = v.strip().strip('"').strip("'")

import httpx
import json
import openpyxl
from wiki_retrieval import build_wiki_context, SKIP_PAGES

# ── Config ────────────────────────────────────────────────────────────────────
WIKI_DIR   = Path(__file__).parent / "wiki"
MODEL      = "gpt-4o"
MAX_TOKENS = 500

EVAL_FILES = [
    {
        "path":       Path("/Users/geodatatek/Downloads/dobyns_rag_evaluated.xlsx"),
        "brand":      "Dobyns Rods",
        "brand_hint": "fishing/dobyns",
    },
    {
        "path":       Path("/Users/geodatatek/Downloads/phenix_rag_evaluated.xlsx"),
        "brand":      "Phenix Rods",
        "brand_hint": "fishing/phenix-rods",
    },
]

SYSTEM_PROMPT = """You are a friendly, knowledgeable customer support assistant for GSM Outdoors.
GSM Outdoors sells hunting and fishing products across multiple brands.

ACCURACY RULES — non-negotiable:
1. Answer ONLY based on the provided wiki content. You may carefully reason about the content, but NEVER use outside knowledge or hallucinate facts not present in the wiki.
2. If a procedure exists in the wiki (warranty steps, cut-and-mail instructions, mailing address, form name, payment amount) — give THOSE EXACT STEPS verbatim. Never substitute "contact customer service" for a real procedure.
3. If a specific model number, part number, address, or price is in the wiki — state it exactly.
4. If the wiki does not contain the answer, say honestly: "I don't have that specific information — please call our support team and they'll be able to help you right away."
4a. HARD RULE — never invent specific facts: If a mailing address, phone number, email address, price, fee, form name, or part number is NOT quoted verbatim in the wiki content provided to you, do NOT write it.
4b. BRAND RULE — never mix brands: Only reference products and details for the confirmed brand.
4c. PRODUCT LINE COMPLETENESS — When a brand has multiple product lines and the customer asks a generic question, address EACH product line separately.

WARRANTY RULES:
9. For ANY warranty or replacement question, ALWAYS include the procedure steps, mailing address, and replacement fee. If the brand has multiple product lines, address the fee for EACH line separately."""


# ── Build wiki cache ──────────────────────────────────────────────────────────
wiki_cache: dict[str, str] = {}
for page in WIKI_DIR.rglob("*.md"):
    rel = str(page.relative_to(WIKI_DIR)).replace("\\", "/")
    if rel not in SKIP_PAGES:
        wiki_cache[rel] = page.read_text()

OPENAI_API_KEY = os.environ["OPENAI_API_KEY"]


# ── Per-question eval ─────────────────────────────────────────────────────────
def run_question(question: str, brand_hint: str, brand: str) -> tuple[str, list[str]]:
    """Return (answer_text, list_of_source_pages)."""
    brand_note = (
        f"⚠️ BRAND CONFIRMED: This conversation is about {brand}. "
        f"Answer every question as if the customer said '{brand}' explicitly.\n\n"
    )
    ctx, ranked = build_wiki_context(question, "", wiki_cache, brand_hint=brand_hint)
    sources = [r for r, _ in ranked] if ranked else []
    messages = [
        {
            "role": "system",
            "content": brand_note + SYSTEM_PROMPT + "\n\n=== WIKI ===\n" + ctx + "\n=== END WIKI ===",
        },
        {"role": "user", "content": question},
    ]
    payload = {"model": MODEL, "messages": messages, "max_tokens": MAX_TOKENS}
    with httpx.Client(timeout=120) as http:
        resp = http.post(
            "https://api.openai.com/v1/chat/completions",
            headers={"Authorization": f"Bearer {OPENAI_API_KEY}", "Content-Type": "application/json"},
            content=json.dumps(payload),
        )
        resp.raise_for_status()
    return resp.json()["choices"][0]["message"]["content"].strip(), sources


def auto_grade(response: str, expected: str) -> str:
    """
    Simple keyword check: split expected into sentences/phrases and check
    if each key number/phrase appears in the response. Returns PASS or FAIL.
    This is a heuristic — review FAIL cases manually.
    """
    resp_lower = response.lower()
    # Extract key numbers and unique phrases from expected answer
    numbers = re.findall(r"\$[\d,]+(?:\.\d+)?|\b\d[\d,]*(?:\.\d+)?\b", expected)
    # Check that key dollar amounts appear in response
    for num in numbers:
        clean = num.replace("$", "").replace(",", "")
        if clean not in resp_lower.replace(",", "") and num.lower() not in resp_lower:
            return "REVIEW"  # numerical mismatch — flag for review
    return "PASS"


def load_questions(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers, row))
        if d.get("Customer Question"):
            rows.append(d)
    return rows


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    total_pass = 0
    total_fail = 0
    all_results = []

    for ef in EVAL_FILES:
        brand     = ef["brand"]
        hint      = ef["brand_hint"]
        questions = load_questions(ef["path"])

        print(f"\n{'='*70}")
        print(f"  {brand}  ({len(questions)} questions)")
        print(f"{'='*70}")

        brand_pass = 0
        brand_fail = 0

        for q_row in questions:
            num      = q_row.get("#") or "?"
            question = q_row.get("Customer Question", "")
            expected = q_row.get("Expected Answer (Key Points)", "") or ""
            old_pf   = q_row.get("Pass / Fail") or "?"

            print(f"\nQ{num}: {question[:75]}", flush=True)

            try:
                answer, sources = run_question(question, hint, brand)
            except Exception as e:
                answer  = f"ERROR: {e}"
                sources = []
            time.sleep(1.5)  # avoid OpenAI rate limits

            grade = auto_grade(answer, expected)

            # Store result
            all_results.append({
                "brand":    brand,
                "num":      num,
                "question": question,
                "expected": expected,
                "answer":   answer,
                "sources":  sources,
                "old_pf":   old_pf,
                "grade":    grade,
            })

            status = "✅" if grade == "PASS" else "🔍 REVIEW"
            print(f"  {status}  Sources: {[s.split('/')[-1].replace('.md','') for s in sources]}")
            if grade != "PASS":
                print(f"  Expected key: {expected[:150]}")
                print(f"  Got: {answer[:200]}")
                brand_fail += 1
            else:
                brand_pass += 1

        print(f"\n  {brand} score: {brand_pass}/{brand_pass+brand_fail}")
        total_pass += brand_pass
        total_fail += brand_fail

    print(f"\n{'='*70}")
    print(f"  TOTAL: {total_pass}/{total_pass+total_fail} passed")
    if total_fail > 0:
        print(f"  {total_fail} questions flagged for review (see above)")
    print(f"{'='*70}\n")

    return all_results


if __name__ == "__main__":
    results = main()
