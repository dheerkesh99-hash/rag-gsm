#!/usr/bin/env python3
"""
generate_rag_eval_excel.py
Run all Dobyns + Bucca questions through the keyword RAG pipeline + gpt-4o-mini,
then write RAG_Evaluated_Dobyns_Bucca.xlsx in the same format as RAG_Evaluated_v5.xlsx.
"""
import sys, time, os
from pathlib import Path

BASE_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(BASE_DIR))

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.series import DataPoint
from openai import OpenAI

from wiki_retrieval import build_wiki_context, merge_page_keywords, SKIP_PAGES
import brand_gate as _brand_gate
from test_rag_dobyns_bucca import DOBYNS_QUESTIONS, BUCCA_QUESTIONS

WIKI_DIR = BASE_DIR / "wiki"

# Same mapping as app_v3._BRAND_WIKI_PREFIX
_BRAND_WIKI_PREFIX = {
    "Phenix Rods":   "fishing/phenix-rods",
    "Dobyns Rods":   "fishing/dobyns-rods",
    "Bucca Brand":   "fishing/bucca-brand",
    "Bonehead Tackle": "fishing/bonehead-tackle",
}
MODEL = "gpt-4o-mini"

# Load API key from .env in same directory
_env_path = BASE_DIR / ".env"
if _env_path.exists():
    for line in _env_path.read_text().splitlines():
        if line.startswith("OPENAI_API_KEY="):
            os.environ.setdefault("OPENAI_API_KEY", line.split("=", 1)[1].strip().strip('"'))

_openai_client = OpenAI(api_key=os.environ["OPENAI_API_KEY"])

SYSTEM_PROMPT = """You are a friendly, knowledgeable customer support assistant for GSM Outdoors.
GSM Outdoors sells hunting and fishing products across multiple brands.

YOUR PERSONALITY:
- Warm and conversational — like a helpful store assistant, not a robot
- Ask natural clarifying questions when needed
- Short sentences, plain language

ACCURACY RULES — non-negotiable:
1. Answer ONLY based on the provided wiki content. You may carefully reason about the content, but NEVER use outside knowledge or hallucinate facts not present in the wiki.
2. If a procedure exists in the wiki (warranty steps, cut-and-mail instructions, mailing address, form name, payment amount) — give THOSE EXACT STEPS verbatim. Never substitute "contact customer service" for a real procedure.
3. If a specific model number, part number, address, or price is in the wiki — state it exactly.
4. If the wiki does not contain the answer, say honestly: "I don't have that specific information — please call our support team and they'll be able to help you right away."

WARRANTY RULES:
5. For ANY warranty or replacement question, ALWAYS include: The procedure steps, the mailing address, and the replacement fee (tier fee + return shipping). Never answer a warranty question without mentioning the cost, if available.

CONVERSATION RULES:
6. After resolving an issue, ask "Is there anything else I can help you with?"
7. Keep answers focused — give the key info first, offer more detail if needed.
"""

# ── Colours ────────────────────────────────────────────────────────────────────
C_TITLE   = "FF1F3864"   # dark navy
C_HEADER  = "FF2E75B6"   # blue
C_PASS    = "FFE2EFDA"   # light green
C_FAIL    = "FFFFC7CE"   # light red
C_PARTIAL = "FFFFEB9C"   # light yellow
C_EXCL    = "FFD9D9D9"   # grey
C_GREY    = "FFF2F2F2"   # light grey
C_WHITE   = "FFFFFFFF"
WHITE_FONT = "FFFFFFFF"
BLACK_FONT = "FF000000"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color=BLACK_FONT, size=10):
    return Font(bold=bold, color=color, size=size)

def thin_border():
    s = Side(border_style="thin", color="FFD9D9D9")
    return Border(left=s, right=s, top=s, bottom=s)

def load_wiki_cache():
    cache = {}
    for page in WIKI_DIR.rglob("*.md"):
        rel = str(page.relative_to(WIKI_DIR)).replace("\\", "/")
        if rel not in SKIP_PAGES:
            cache[rel] = page.read_text()
    return cache

def get_context_and_sources(question: str, wiki_cache: dict) -> tuple[str, str, float, str]:
    """Return (context_text, formatted_sources, top_score, detected_brand).
    Mirrors app_v3: run brand_gate first, pass brand_hint to build_wiki_context.
    """
    gate_result = _brand_gate.detect_brand(question, [], {})
    brand_hint = None
    detected_brand = ""
    if gate_result.get("brands"):
        detected_brand = gate_result["brands"][0]
        brand_hint = _BRAND_WIKI_PREFIX.get(detected_brand)

    ctx, sources = build_wiki_context(question, "", wiki_cache, brand_hint=brand_hint)
    if not sources:
        return ctx, "No sources retrieved", 0.0, detected_brand
    fmt = "\n".join(
        f"{rel.upper()} (Match Score: {score:.4f})"
        for rel, score in sources[:5]
    )
    top = sources[0][1] if sources else 0.0
    return ctx, fmt, top, detected_brand

def ask_llm(question: str, context: str) -> tuple[str, float, str]:
    """Call gpt-4o-mini via OpenAI. Returns (response_text, latency_sec, note_suffix)."""
    full_system = (
        SYSTEM_PROMPT
        + "\n\n=== GSM OUTDOORS PRODUCT WIKI ===\n"
        + "Use ONLY the information below to answer customer questions. "
        + "If the answer is not here, say so honestly.\n\n"
        + context
        + "\n=== END WIKI ==="
    )
    t0 = time.time()
    try:
        resp = _openai_client.chat.completions.create(
            model=MODEL,
            messages=[
                {"role": "system", "content": full_system},
                {"role": "user",   "content": question},
            ],
            timeout=60,
        )
        text = resp.choices[0].message.content.strip()
        in_t  = resp.usage.prompt_tokens
        out_t = resp.usage.completion_tokens
        latency = time.time() - t0
        note_suffix = f"[{MODEL} | In:{in_t} Out:{out_t} tok]"
        return text, latency, note_suffix
    except Exception as e:
        return f"ERROR: {e}", time.time() - t0, ""

def source_file_short(expected: str) -> str:
    return Path(expected).name

def add_pie_chart(ws, total_rows: int):
    """Add a Pass/Fail pie chart to the worksheet using a hidden summary at column P."""
    pass_count = sum(
        1 for r in range(4, 4 + total_rows)
        if "PASS" in str(ws.cell(row=r, column=9).value or "")
    )
    fail_count = total_rows - pass_count

    # Write summary data at col P/Q (cols 16/17) — used as chart source
    for r, (label, count) in enumerate(
        [("Result", "Count"), ("Pass", pass_count), ("Fail", fail_count)], start=1
    ):
        ws.cell(row=r, column=16, value=label)
        ws.cell(row=r, column=17, value=count)

    chart = PieChart()
    chart.title = f"{ws.title} — Pass / Fail"
    chart.style = 10
    chart.width  = 14
    chart.height = 10

    data   = Reference(ws, min_col=17, min_row=1, max_row=3)
    labels = Reference(ws, min_col=16, min_row=2, max_row=3)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)

    # Colour slices: index 0 = Pass (green), index 1 = Fail (red)
    series = chart.series[0]
    for idx, color in enumerate(["00B050", "FF0000"]):
        pt = DataPoint(idx=idx)
        pt.graphicalProperties.solidFill = color
        series.dPt.append(pt)

    ws.add_chart(chart, "P5")

def build_sheet(ws, brand: str, questions: list, wiki_cache: dict):
    """Populate one worksheet."""
    ws.title = brand

    # ── Row 1: Title bar ──────────────────────────────────────────────────────
    ws.merge_cells("A1:N1")
    c = ws["A1"]
    c.value = f"RAG SYSTEM TEST LOG — {brand.upper()}"
    c.fill = fill(C_TITLE)
    c.font = font(bold=True, color=WHITE_FONT, size=12)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # ── Row 2: Pass criteria ──────────────────────────────────────────────────
    ws.merge_cells("A2:N2")
    c = ws["A2"]
    c.value = "Pass Criteria: Correct answer + right source page in Retrieval Sources"
    c.font = font(bold=False, color=BLACK_FONT, size=9)
    c.alignment = Alignment(horizontal="left", vertical="center")

    # ── Row 3: Headers ────────────────────────────────────────────────────────
    headers = ["#", "Source File", "Question", "Expected Answer Summary",
               "Actual Response (paste)", "Retrieval Sources Returned",
               "Correct\nAnswer?", "Correct\nSource?", "Pass / Fail",
               "Latency\n(sec)", "Top Retrieval\nScore",
               "Hallucination\nDetected?", "Notes / Observations", "Tester"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=ci, value=h)
        c.fill = fill(C_HEADER)
        c.font = font(bold=True, color=WHITE_FONT, size=9)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[3].height = 30

    # ── Column widths ─────────────────────────────────────────────────────────
    widths = [4, 28, 45, 38, 45, 38, 9, 9, 10, 8, 11, 12, 52, 12]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # ── Data rows ─────────────────────────────────────────────────────────────
    total = len(questions)
    for idx, (qnum, expected_src, baseline, question) in enumerate(questions):
        row = idx + 4
        print(f"  [{brand}] Q{qnum}/{total}: {question[:55]}...", flush=True)

        ctx, sources_str, top_score, detected_brand = get_context_and_sources(question, wiki_cache)
        response, latency, note_suffix = ask_llm(question, ctx)
        if detected_brand:
            note_suffix = f"Brand: {detected_brand} | {note_suffix}"

        # Determine if expected source was retrieved
        src_in_sources = expected_src.upper().replace("/", "/") in sources_str.upper()

        # Pass/fail cell value + colour
        if src_in_sources:
            pf_val   = "✅ PASS"
            pf_color = C_PASS
        else:
            pf_val   = "❌ FAIL"
            pf_color = C_FAIL

        row_data = [
            qnum,                          # A: #
            source_file_short(expected_src),  # B: Source File
            question,                      # C: Question
            "",                            # D: Expected Answer Summary (blank — to fill)
            response,                      # E: Actual Response
            sources_str,                   # F: Retrieval Sources
            "",                            # G: Correct Answer?
            "Yes" if src_in_sources else "No",  # H: Correct Source?
            pf_val,                        # I: Pass/Fail
            round(latency, 2),             # J: Latency
            round(top_score, 4),           # K: Top Score
            "",                            # L: Hallucination
            note_suffix,                   # M: Notes
            "",                            # N: Tester
        ]

        for ci, val in enumerate(row_data, 1):
            c = ws.cell(row=row, column=ci, value=val)
            c.border = thin_border()
            c.alignment = Alignment(wrap_text=True, vertical="top")
            c.font = font(size=9)

            # Column-specific fills
            if ci in (4, 5, 6, 10, 11, 13, 14):
                c.fill = fill(C_GREY)
            elif ci in (7, 8, 9, 12):
                c.fill = fill(pf_color)
            else:
                c.fill = fill(C_WHITE)

        ws.row_dimensions[row].height = 60

    # ── Pie chart ─────────────────────────────────────────────────────────────
    add_pie_chart(ws, total)

    # ── Freeze panes ──────────────────────────────────────────────────────────
    ws.freeze_panes = "A4"

def main():
    out_path = BASE_DIR / "RAG_Evaluated_Dobyns_Bucca.xlsx"
    print("Loading wiki cache...")
    wiki_cache = load_wiki_cache()
    print(f"  {len(wiki_cache)} pages loaded")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    print(f"\nProcessing {len(DOBYNS_QUESTIONS)} Dobyns questions...")
    ws_d = wb.create_sheet("Dobyns Rods")
    build_sheet(ws_d, "Dobyns Rods", DOBYNS_QUESTIONS, wiki_cache)

    print(f"\nProcessing {len(BUCCA_QUESTIONS)} Bucca Brand questions...")
    ws_b = wb.create_sheet("Bucca Brand")
    build_sheet(ws_b, "Bucca Brand", BUCCA_QUESTIONS, wiki_cache)

    wb.save(out_path)
    print(f"\n✅ Saved: {out_path}")

if __name__ == "__main__":
    main()
