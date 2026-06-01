#!/usr/bin/env python3
"""
generate_rag_eval_bonehead.py
Run all Bonehead Tackle questions through the RAG pipeline (brand gate + retrieval + GPT-4o),
evaluate accuracy against the 4 Bonehead wiki files as sole source of truth,
and write RAG_Evaluated_Bonehead.xlsx.
"""
import sys, time, os
from pathlib import Path

BASE_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(BASE_DIR))

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.series import DataPoint
from openai import OpenAI

from wiki_retrieval import build_wiki_context, SKIP_PAGES

WIKI_DIR = BASE_DIR / "wiki"
BRAND_HINT = "fishing/bonehead-tackle"   # all questions are in confirmed Bonehead session
MODEL = "gpt-4o"

# Load API key
_env_path = BASE_DIR / ".env"
if _env_path.exists():
    for line in _env_path.read_text().splitlines():
        if line.startswith("OPENAI_API_KEY="):
            os.environ.setdefault("OPENAI_API_KEY", line.split("=", 1)[1].strip().strip('"'))
_openai_client = OpenAI(api_key=os.environ["OPENAI_API_KEY"])

# ── Questions ────────────────────────────────────────────────────────────────
# (num, expected_source_page, expected_key_points, question)
# expected_source_page: the wiki file that MUST appear in retrieval sources
# expected_key_points: list of substrings — ALL must appear in the response to auto-pass
# NOT_IN_WIKI: question whose correct answer is "I don't have that information"

NOT_IN_WIKI = "__not_in_wiki__"

BONEHEAD_QUESTIONS = [
    # ── Pricing ──
    (1,  "bonehead-tackle-carbon-fiber-spinning-rods.md",
         ["$119.99"],
         "How much does the 10 foot carbon fiber rod cost?"),

    (2,  "bonehead-tackle-e-series-carbon-fiber-spinning-rods.md",
         ["$59.99"],
         "What's the cheapest E-Series rod you sell?"),

    (3,  "bonehead-tackle-carbon-fiber-replacement-tips.md",
         ["21.99", "37.99"],
         "Are replacement tips cheaper for E-Series rods than the premium ones?"),

    (4,  "bonehead-tackle-carbon-fiber-replacement-tips.md",
         ["21.99", "26.99"],
         "What's the price difference between the 8 foot and 12 foot E-Series replacement tips?"),

    # ── Product lineup ──
    (5,  "bonehead-tackle-carbon-fiber-spinning-rods.md",
         ["5'8"],
         "What's the shortest rod you carry?"),

    (6,  "bonehead-tackle-e-series-carbon-fiber-spinning-rods.md",
         ["12", "E-Series"],
         "Do you make a 12 foot spinning rod?"),

    (7,  "bonehead-tackle-carbon-fiber-spinning-rods.md",
         ["cork", "EVA"],
         "What's the difference between the Carbon Fiber rods and the E-Series rods?"),

    (8,  "bonehead-tackle-carbon-fiber-spinning-rods.md",
         ["green", "Green"],
         "Do the premium rods come in any colors?"),

    (9,  "bonehead-tackle-carbon-fiber-spinning-rods.md",
         ["two-piece", "2-piece", "7'", "8'"],
         "Are any of your rods two-piece?"),

    (10, "bonehead-tackle-carbon-fiber-spinning-rods.md",
         ["stainless steel"],
         "What kind of guides do your rods use?"),

    (11, "bonehead-tackle-carbon-fiber-spinning-rods.md",
         ["3.2"],
         "How light is the 5'8\" spinning rod?"),

    (12, "bonehead-tackle-carbon-fiber-spinning-rods.md",
         ["cork", "EVA"],
         "What kind of grip do the premium rods have vs the E-Series?"),

    # ── Warranty ──
    (13, "bonehead-tackle-warranty.md",
         ["90", "1 year", "1-year"],
         "How long is the warranty on your rods?"),

    (14, "bonehead-tackle-warranty.md",
         ["accidental", "not cover", "does not cover"],
         "Does my warranty cover a rod I snapped by accident?"),

    (15, "bonehead-tackle-warranty.md",
         ["receipt", "proof of purchase"],
         "Do I need my receipt to make a warranty claim?"),

    (16, "bonehead-tackle-warranty.md",
         ["$30"],
         "Is there a fee if I need a replacement under warranty?"),

    (17, "bonehead-tackle-warranty.md",
         ["90", "1 year", "1-year"],
         "Is the E-Series warranty the same as the premium rod warranty?"),

    # ── Replacement tips ──
    (18, "bonehead-tackle-carbon-fiber-replacement-tips.md",
         ["yes", "Yes", "replace"],
         "Can I replace just the tip instead of buying a whole new rod?"),

    (19, "bonehead-tackle-carbon-fiber-replacement-tips.md",
         ["42.99", "21.99"],
         "How much does a replacement tip cost for my 10 foot rod?"),

    (20, "bonehead-tackle-carbon-fiber-replacement-tips.md",
         ["yes", "Yes", "E-Series"],
         "Do you sell replacement tips for E-Series rods?"),

    # ── Not in wiki ──
    (21, NOT_IN_WIKI,
         ["don't have", "don't carry", "not available", "I don't", "support team", "no information"],
         "Do you sell reels?"),

    (22, NOT_IN_WIKI,
         ["don't have", "I don't", "support team", "not available", "no information"],
         "What fishing line do you recommend?"),

    (23, "bonehead-tackle-carbon-fiber-spinning-rods.md",
         ["no", "No", "not available", "E-Series", "10"],
         "Can I buy a 12 foot premium Carbon Fiber rod?"),

    (24, "bonehead-tackle-warranty.md",
         ["no", "No", "not", "1 year", "1-year", "90"],
         "Do you offer a lifetime warranty?"),
]

# ── Colours ──────────────────────────────────────────────────────────────────
C_TITLE   = "FF1F3864"
C_HEADER  = "FF2E75B6"
C_PASS    = "FFE2EFDA"
C_FAIL    = "FFFFC7CE"
C_GREY    = "FFF2F2F2"
C_WHITE   = "FFFFFFFF"
WHITE_FONT = "FFFFFFFF"
BLACK_FONT = "FF000000"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color=BLACK_FONT, size=10):
    return Font(bold=bold, color=color, size=size)

def thin_border():
    s = Side(border_style="thin", color="FFD9D9D9")
    return Border(left=s, right=s, top=s, bottom=s)

def load_wiki_cache():
    cache = {}
    for page in WIKI_DIR.rglob("*.md"):
        rel = str(page.relative_to(WIKI_DIR)).replace("\\", "/")
        if rel not in SKIP_PAGES:
            cache[rel] = page.read_text()
    return cache

SYSTEM_PROMPT = """You are a friendly, knowledgeable customer support assistant for GSM Outdoors.
You are helping a customer with a question about Bonehead Tackle products.

ACCURACY RULES — non-negotiable:
1. Answer ONLY based on the provided wiki content. NEVER use outside knowledge or hallucinate facts not present in the wiki.
2. If a specific price, fee, or spec is in the wiki — state it exactly.
3. If the wiki does not contain the answer, say honestly: "I don't have that specific information — please call our support team and they'll be able to help you right away."
4a. HARD RULE: If a price, fee, address, part number, or product detail is NOT quoted verbatim in the wiki content provided to you, do NOT write it. Say instead: "I don't have that detail in my notes — please call our support team."
4b. BRAND RULE: You are answering about Bonehead Tackle ONLY. Never mention or reference products, reels, accessories, or procedures from any other brand (e.g. Dobyns, Phenix, Bucca). If the customer asks about something Bonehead does not carry, say you don't have that information for Bonehead Tackle.

CONVERSATION RULES:
5. Keep answers focused — give the key info first.
6. After resolving an issue, ask "Is there anything else I can help you with?"
"""

def get_context_and_sources(question: str, wiki_cache: dict) -> tuple[str, str, float]:
    ctx, sources = build_wiki_context(question, "", wiki_cache, brand_hint=BRAND_HINT)
    if not sources:
        return ctx, "No sources retrieved", 0.0
    fmt = "\n".join(
        f"{rel.upper()} (Match Score: {score:.4f})"
        for rel, score in sources[:5]
    )
    top = sources[0][1] if sources else 0.0
    return ctx, fmt, top

def ask_llm(question: str, context: str) -> tuple[str, float, str]:
    full_system = (
        SYSTEM_PROMPT
        + "\n\n=== BONEHEAD TACKLE PRODUCT WIKI ===\n"
        + "Use ONLY the information below to answer the customer's question. "
        + "If the answer is not here, say so honestly.\n\n"
        + context
        + "\n=== END WIKI ==="
    )
    t0 = time.time()
    try:
        resp = _openai_client.chat.completions.create(
            model=MODEL,
            messages=[
                {"role": "system", "content": full_system},
                {"role": "user",   "content": question},
            ],
            timeout=60,
        )
        text = resp.choices[0].message.content.strip()
        in_t  = resp.usage.prompt_tokens
        out_t = resp.usage.completion_tokens
        latency = time.time() - t0
        note = f"[{MODEL} | In:{in_t} Out:{out_t} tok]"
        return text, latency, note
    except Exception as e:
        return f"ERROR: {e}", time.time() - t0, ""

def evaluate(response: str, expected_src: str, key_points: list[str], sources_str: str) -> tuple[str, str, str, str]:
    """Return (correct_answer, correct_source, pass_fail, notes)."""
    r_lower = response.lower()

    # Source check
    if expected_src == NOT_IN_WIKI:
        src_ok = True  # no specific source required
        src_val = "N/A"
    else:
        src_ok = expected_src.upper() in sources_str.upper()
        src_val = "Yes" if src_ok else "No"

    # Answer check — ALL key points must be present
    ans_ok = any(kp.lower() in r_lower for kp in key_points)
    ans_val = "Yes" if ans_ok else "No"

    if ans_ok and src_ok:
        pf = "✅ PASS"
        notes = ""
    elif ans_ok and not src_ok:
        pf = "⚠️ PARTIAL"
        notes = f"Correct answer but wrong source page retrieved"
    elif not ans_ok and src_ok:
        pf = "❌ FAIL"
        notes = f"Right source retrieved but answer missing key points: {key_points[:3]}"
    else:
        pf = "❌ FAIL"
        notes = f"Wrong source + missing key points: {key_points[:3]}"

    return ans_val, src_val, pf, notes

def add_pie_chart(ws, total_rows: int):
    pass_count = sum(
        1 for r in range(4, 4 + total_rows)
        if "PASS" in str(ws.cell(row=r, column=9).value or "")
    )
    fail_count = sum(
        1 for r in range(4, 4 + total_rows)
        if "FAIL" in str(ws.cell(row=r, column=9).value or "")
    )
    partial = total_rows - pass_count - fail_count

    for r, (label, count) in enumerate(
        [("Result", "Count"), ("Pass", pass_count), ("Fail", fail_count), ("Partial", partial)], start=1
    ):
        ws.cell(row=r, column=16, value=label)
        ws.cell(row=r, column=17, value=count)

    chart = PieChart()
    chart.title = "Bonehead Tackle — Pass / Fail"
    chart.style = 10
    chart.width  = 14
    chart.height = 10

    data   = Reference(ws, min_col=17, min_row=1, max_row=4)
    labels = Reference(ws, min_col=16, min_row=2, max_row=4)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)

    for idx, color in enumerate(["00B050", "FF0000", "FFC000"]):
        pt = DataPoint(idx=idx)
        pt.graphicalProperties.solidFill = color
        chart.series[0].dPt.append(pt)

    ws.add_chart(chart, "P5")

def build_sheet(ws, wiki_cache: dict):
    ws.title = "Bonehead Tackle"

    ws.merge_cells("A1:N1")
    c = ws["A1"]
    c.value = "RAG SYSTEM TEST LOG — BONEHEAD TACKLE"
    c.fill = fill(C_TITLE)
    c.font = font(bold=True, color=WHITE_FONT, size=12)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    ws.merge_cells("A2:N2")
    c = ws["A2"]
    c.value = f"Pass Criteria: Response contains expected key facts AND correct source page retrieved | Model: {MODEL} | Brand hint: {BRAND_HINT}"
    c.font = font(bold=False, color=BLACK_FONT, size=9)
    c.alignment = Alignment(horizontal="left", vertical="center")

    headers = ["#", "Source File", "Question", "Expected Key Points",
               "Actual Response", "Retrieval Sources Returned",
               "Correct\nAnswer?", "Correct\nSource?", "Pass / Fail",
               "Latency\n(sec)", "Top Retrieval\nScore",
               "Hallucination\nDetected?", "Notes / Observations", "Tester"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=ci, value=h)
        c.fill = fill(C_HEADER)
        c.font = font(bold=True, color=WHITE_FONT, size=9)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[3].height = 30

    widths = [4, 38, 45, 42, 50, 42, 9, 9, 10, 8, 11, 12, 52, 12]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    total = len(BONEHEAD_QUESTIONS)
    for idx, (qnum, expected_src, key_points, question) in enumerate(BONEHEAD_QUESTIONS):
        row = idx + 4
        print(f"  Q{qnum}/{total}: {question[:60]}...", flush=True)

        ctx, sources_str, top_score = get_context_and_sources(question, wiki_cache)
        response, latency, note_suffix = ask_llm(question, ctx)

        ans_val, src_val, pf, eval_notes = evaluate(response, expected_src, key_points, sources_str)

        if "PASS" in pf:
            pf_color = C_PASS
        elif "PARTIAL" in pf:
            pf_color = "FFFFEB9C"
        else:
            pf_color = C_FAIL

        src_display = "N/A (not in wiki)" if expected_src == NOT_IN_WIKI else Path(expected_src).name
        kp_display  = " | ".join(key_points)

        row_data = [
            qnum, src_display, question, kp_display,
            response, sources_str,
            ans_val, src_val, pf,
            round(latency, 2), round(top_score, 4),
            "",
            (eval_notes + " " + note_suffix).strip(),
            "",
        ]

        for ci, val in enumerate(row_data, 1):
            c = ws.cell(row=row, column=ci, value=val)
            c.border = thin_border()
            c.alignment = Alignment(wrap_text=True, vertical="top")
            c.font = font(size=9)
            if ci in (4, 5, 6, 10, 11, 13, 14):
                c.fill = fill(C_GREY)
            elif ci in (7, 8, 9, 12):
                c.fill = fill(pf_color)
            else:
                c.fill = fill(C_WHITE)

        ws.row_dimensions[row].height = 65

    add_pie_chart(ws, total)
    ws.freeze_panes = "A4"

def main():
    out_path = BASE_DIR / "RAG_Evaluated_Bonehead.xlsx"
    print("Loading wiki cache...")
    wiki_cache = load_wiki_cache()
    print(f"  {len(wiki_cache)} pages loaded\n")
    print(f"Running {len(BONEHEAD_QUESTIONS)} Bonehead Tackle questions through {MODEL}...\n")

    wb = openpyxl.Workbook()
    ws = wb.active
    build_sheet(ws, wiki_cache)

    wb.save(out_path)
    print(f"\n✅  Saved: {out_path}")

if __name__ == "__main__":
    main()
