#!/usr/bin/env python3
"""
add_charts_to_eval.py
Post-processor: adds Pass/Fail pie charts to an existing RAG eval Excel file.
Usage: python3 add_charts_to_eval.py [path/to/file.xlsx]
"""
import sys
from pathlib import Path
import openpyxl
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.series import DataPoint

def add_pie_chart(ws):
    total_rows = ws.max_row - 3  # rows 1-3 are title/criteria/headers
    if total_rows <= 0:
        return

    pass_count = sum(
        1 for r in range(4, 4 + total_rows)
        if "PASS" in str(ws.cell(row=r, column=9).value or "")
    )
    fail_count = total_rows - pass_count

    # Write summary data at col P/Q (cols 16/17)
    for r, (label, count) in enumerate(
        [("Result", "Count"), ("Pass", pass_count), ("Fail", fail_count)], start=1
    ):
        ws.cell(row=r, column=16, value=label)
        ws.cell(row=r, column=17, value=count)

    chart = PieChart()
    chart.title = f"{ws.title} — Pass / Fail"
    chart.style = 10
    chart.width  = 14
    chart.height = 10

    data   = Reference(ws, min_col=17, min_row=1, max_row=3)
    labels = Reference(ws, min_col=16, min_row=2, max_row=3)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)

    series = chart.series[0]
    for idx, color in enumerate(["00B050", "FF0000"]):  # green=Pass, red=Fail
        pt = DataPoint(idx=idx)
        pt.graphicalProperties.solidFill = color
        series.dPt.append(pt)

    ws.add_chart(chart, "P5")
    print(f"  {ws.title}: {pass_count} Pass / {fail_count} Fail")

def main():
    path = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("RAG_Evaluated_Dobyns_Bucca.xlsx")
    if not path.exists():
        print(f"File not found: {path}")
        sys.exit(1)

    wb = openpyxl.load_workbook(path)
    for ws in wb.worksheets:
        print(f"Adding chart to sheet: {ws.title}")
        add_pie_chart(ws)

    wb.save(path)
    print(f"\n✅ Charts added: {path}")

if __name__ == "__main__":
    main()
